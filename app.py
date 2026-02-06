import streamlit as st
import pandas as pd
import zipfile
import os
import re
import shutil
import docx
from pypdf import PdfReader
from io import BytesIO
import extract_msg
from email import policy
from email.parser import BytesParser
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ==========================================
# 🧠 SENIOR DEV INTELLIGENCE: CONCEPT MAPPING
# ==========================================
# Yahan hum Script ko "Samajh" (Context) de rahe hain.
# LEFT: Excel Column Name
# RIGHT: List of keywords/phrases jo email mein ho sakte hain

SMART_MAPPINGS = {
    "UAT signoff by": [
        "tested by", "checked by", "validated by", "approved by", 
        "signoff from", "confirmed by", "regards", "thanks"
    ],
    "UAT signoff on": [
        "date", "dated", "on", "completed on", "done on", "timestamp"
    ],
    "Change description": [
        "regarding", "change for", "deployment of", "issue with", 
        "summary", "requirement"
    ],
    "Downtime Approval": [
        "downtime approved", "window provided", "maintenance window", 
        "shutdown allowed", "service disruption"
    ],
    "Implementation team approval by": [
        "implementation approved", "deploy it", "proceed with", 
        "good to go", "go ahead"
    ]
}

# Agar data mil jaye, to kya hum poori line utha lein? (Evidence ke liye better hai)
CAPTURE_WHOLE_LINE = True 

# ==========================================
# FILE READERS (STANDARD)
# ==========================================
def get_file_text(file_path):
    """Detects file type and extracts clean text line-by-line."""
    text = ""
    ext = file_path.lower()
    
    try:
        if ext.endswith('.pdf'):
            reader = PdfReader(file_path)
            for page in reader.pages:
                text += page.extract_text() + "\n"
        
        elif ext.endswith('.docx'):
            doc = docx.Document(file_path)
            text = "\n".join([p.text for p in doc.paragraphs])
            
        elif ext.endswith('.msg'):
            msg = extract_msg.Message(file_path)
            # Subject is often critical for context
            text = f"Subject: {msg.subject}\n{msg.body}"
            
        elif ext.endswith('.eml'):
            with open(file_path, 'rb') as f:
                msg = BytesParser(policy=policy.default).parse(f)
                text = f"Subject: {msg['subject']}\n{msg.get_body(preferencelist=('plain')).get_content()}"
    except:
        return ""
    
    return text

# ==========================================
# 🕵️ LOGIC: CONTEXTUAL FINDER
# ==========================================
def find_answer_in_context(column_name, full_text):
    """
    Excel Header ke keywords ko Email Text mein dhoondta hai.
    Regex nahi, ye 'Meaning' dhoond raha hai.
    """
    if not full_text: return None
    
    # Text ko lines mein todo taaki hum specific line pakad sakein
    lines = full_text.split('\n')
    
    # Check agar column hamari mapping dictionary mein hai
    # Agar nahi hai, to Column Name ko hi keyword maan lo
    keywords = SMART_MAPPINGS.get(column_name, [column_name])
    
    best_match = None
    
    for line in lines:
        line_clean = line.strip()
        if not line_clean: continue
        
        # Har keyword check karo
        for keyword in keywords:
            # Case Insensitive Search
            if keyword.lower() in line_clean.lower():
                
                # Agar keyword mil gaya, to us line ka RELEVANT hissa nikalo
                
                # Logic 1: Agar line keyword se start hoti hai (e.g., "Approved by: Rahul")
                # To colon (:) ya hyphen (-) ke baad ka text uthao
                if ":" in line_clean:
                    parts = line_clean.split(":", 1)
                    if keyword.lower() in parts[0].lower(): # Keyword left side pe hai
                        return parts[1].strip()
                
                # Logic 2: Contextual Inference (Senior Logic)
                # Agar likha hai "Please proceed with deployment - Rahul"
                # Aur keyword "proceed" hai, to poori line hi answer hai.
                return line_clean # Poori line return kardo safe side ke liye
                
    return None

# ==========================================
# MAIN EXECUTION
# ==========================================
def run_senior_audit(excel_file, zip_file, id_col):
    # 1. Setup
    temp_dir = "temp_audit_ai"
    if os.path.exists(temp_dir): shutil.rmtree(temp_dir)
    os.makedirs(temp_dir)
    
    with zipfile.ZipFile(zip_file, 'r') as z:
        z.extractall(temp_dir)

    all_files = []
    for root, _, files in os.walk(temp_dir):
        for f in files:
            all_files.append(os.path.join(root, f))

    # 2. Excel Setup
    df = pd.read_excel(excel_file)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
    output.seek(0)
    
    wb = load_workbook(output)
    ws = wb.active
    
    headers = [str(c.value).strip() for c in ws[1]]
    
    # Styles
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    if id_col not in headers:
        return None

    # 3. Processing
    bar = st.progress(0)
    logs = st.empty()
    
    for i, row in df.iterrows():
        bar.progress((i + 1) / len(df))
        
        row_id = str(row.get(id_col, "")).strip().lower()
        if not row_id or row_id == "nan": continue
        
        # Find File (Fuzzy Match ID in Filename)
        file_text = None
        for f in all_files:
            if row_id in os.path.basename(f).lower():
                file_text = get_file_text(f)
                break
        
        # Row Processing
        excel_row = i + 2
        
        if file_text:
            # Har column ke liye scan karo
            for col_idx, col_name in enumerate(headers):
                if col_name == id_col: continue # ID ko skip karo
                
                cell = ws.cell(row=excel_row, column=col_idx+1)
                
                # Agar cell pehle se bhara hai to chhod do
                if cell.value: continue
                
                # --- THE BRAIN ---
                found_value = find_answer_in_context(col_name, file_text)
                
                if found_value:
                    # Value mil gayi!
                    # Thoda sa safai (Cleaning)
                    clean_val = found_value.replace("\t", " ").strip()
                    cell.value = clean_val
                    cell.fill = green_fill
                else:
                    # File mili par data samajh nahi aaya -> Yellow
                    # Isse aap manually check kar paoge
                    cell.fill = yellow_fill
        else:
            # File hi nahi mili -> Poori row Yellow
            for c in range(1, len(headers) + 1):
                ws.cell(row=excel_row, column=c).fill = yellow_fill

    # Finalize
    shutil.rmtree(temp_dir)
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out

# ==========================================
# UI LAYOUT
# ==========================================
st.title("🧠 Context-Aware Audit Bot")
st.markdown("""
Ye bot **'Keywords'** aur **'Context'** padhta hai, sirf labels nahi.
Agar Excel mein **'Approver'** manga hai, ye email mein **'Regards'** ya **'Thanks'** dhoond lega.
""")

c1, c2 = st.columns(2)
up_xl = c1.file_uploader("Upload Excel", type=["xlsx"])
up_zip = c2.file_uploader("Upload Emails Zip", type=["zip"])

if up_xl and up_zip:
    df_p = pd.read_excel(up_xl)
    id_s = st.selectbox("Unique ID Column select karein:", df_p.columns)
    
    if st.button("🧠 Start Smart Analysis"):
        with st.spinner("Reading emails and understanding context..."):
            res = run_senior_audit(up_xl, up_zip, id_s)
            if res:
                st.success("Analysis Complete!")
                st.download_button("📥 Download Smart Filled Excel", res, "Smart_Audit_Output.xlsx")
